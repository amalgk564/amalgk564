import pandas as pd
import openpyxl as py
import tkinter as tk
from tkinter import filedialog
from tkinter import *
from tkinter import simpledialog

root = Tk()
SD = filedialog.askdirectory(title="Select folder to save output file")
DTA = filedialog.askopenfilename(initialdir="/", title="Mix Actuals - Export-Non Pro", filetypes=(("allfiles", "*.xlsx"), ("allfiles", "*.*")))
output_file = tk.filedialog.askopenfilename(initialdir="/", title="Select output file", filetypes=(("allfiles", "*.xlsx"), ("allfiles", "*.*")))

root.destroy()

def get_Current_Month1():
    new_col_name = simpledialog.askstring(title="Actuals Month", prompt="Enter the Current Month:")
    return new_col_name

root = tk.Tk()
root.withdraw()

Current_Month1 = get_Current_Month1()


Export_Mix_df = pd.read_excel(DTA, header=[0], sheet_name='Summary')
Export_Mix_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
Export_Mix_Load_df = Export_Mix_df[['Vehicle Line', 'C122 Export Mix','C424 FAP Export Mix','C424 FoE Export Mix','C424 FSA Export Mix','C424 MEA Export Mix']]



Export_Mix_Load_df


OP=py.load_workbook(output_file)

Sheet1 = OP.get_sheet_by_name('C122 EXPORT MIX')
Sheet2 = OP.get_sheet_by_name('C424 EXPORT MIX - FAP')
Sheet3 = OP.get_sheet_by_name('C424 EXPORT MIX - FOE')
Sheet4 = OP.get_sheet_by_name('C424 EXPORT MIX - FSA')
Sheet5 = OP.get_sheet_by_name('C424 EXPORT MIX - MEA')



if Current_Month1 == "January":
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=3)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=3)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=3)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=3)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=3)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'February':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=5)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=5)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=5)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=5)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=5)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'March':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=7)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=7)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=7)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=7)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=7)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'April':
	    for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=11)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=11)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=11)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=11)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=11)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'May':
	    for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=13)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=13)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=13)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=13)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=13)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'June':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=15)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=15)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=15)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=15)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=15)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'July':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=19)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=19)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=19)
            cell3.value = Export_Mix_Load_df.iloc[i-9,3]
            cell4 = Sheet4.cell(row=i, column=19)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=19)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'August':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=21)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=21)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=21)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=21)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=21)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'September':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=23)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=23)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=23)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=23)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=23)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'October':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=27)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=27)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=27)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=27)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=27)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
elif Current_Month1 == 'November':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=29)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=29)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=29)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=29)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=29)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]
else:
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=31)
            cell1.value = Export_Mix_Load_df.iloc[i-9, 1]
            cell2 = Sheet2.cell(row=i, column=31)
            cell2.value = Export_Mix_Load_df.iloc[i-9, 2]
            cell3 = Sheet3.cell(row=i, column=31)
            cell3.value = Export_Mix_Load_df.iloc[i-9, 3]
            cell4 = Sheet4.cell(row=i, column=31)
            cell4.value = Export_Mix_Load_df.iloc[i-9, 4]
            cell5 = Sheet5.cell(row=i, column=31)
            cell5.value = Export_Mix_Load_df.iloc[i-9, 5]

OP.save(SD+"\Exchange.xlsx")