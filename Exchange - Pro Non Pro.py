import pandas as pd
import openpyxl as py
import tkinter as tk
from tkinter import filedialog
from tkinter import *
from tkinter import simpledialog

root = Tk()
SD = filedialog.askdirectory(title="Select folder to save output file")
DTA = filedialog.askopenfilename(initialdir="/", title="Forecast Control Submission", filetypes=(("allfiles", "*.xlsx"), ("allfiles", "*.*")))
output_file = tk.filedialog.askopenfilename(initialdir="/", title="Select output file", filetypes=(("allfiles", "*.xlsx"), ("allfiles", "*.*")))

root.destroy()

def get_Current_Month1():
    new_col_name = simpledialog.askstring(title="Actuals Month", prompt="Enter the Current Month:")
    return new_col_name

root = tk.Tk()
root.withdraw()

Current_Month1 = get_Current_Month1()


EXCHANGE_USA_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE USA')
EXCHANGE_USA_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_USA_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
USAload_df = EXCHANGE_USA_df[['Vehicle Line', Current_Month1]]


EXCHANGE_Canada_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE Canada')
EXCHANGE_Canada_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_Canada_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
Canadaload_df = EXCHANGE_Canada_df[['Vehicle Line', Current_Month1]]


EXCHANGE_Mexico_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE Mexico')
EXCHANGE_Mexico_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_Mexico_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
Mexicoload_df = EXCHANGE_Mexico_df[['Vehicle Line', Current_Month1]]


EXCHANGE_DM_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE DM')
EXCHANGE_DM_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_DM_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
DMload_df = EXCHANGE_DM_df[['Vehicle Line', Current_Month1]]




EXCHANGE_FAP_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE FAP')
EXCHANGE_FAP_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_FAP_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
FAPload_df = EXCHANGE_FAP_df[['Vehicle Line', Current_Month1]]


EXCHANGE_FoE_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE FoE')
EXCHANGE_FoE_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_FoE_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
FoEload_df = EXCHANGE_FoE_df[['Vehicle Line', Current_Month1]]


EXCHANGE_FSA_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE FSA')
EXCHANGE_FSA_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_FSA_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
FSAload_df = EXCHANGE_FSA_df[['Vehicle Line', Current_Month1]]



EXCHANGE_MEA_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE MEA')
EXCHANGE_MEA_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_MEA_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
MEAload_df = EXCHANGE_MEA_df[['Vehicle Line', Current_Month1]]




EXCHANGE_FNA_df = pd.read_excel(DTA, header=[5], sheet_name='EXCHANGE FNA')
EXCHANGE_FNA_df.rename(columns={'Unnamed: 0': 'Vehicle Line'}, inplace=True)
EXCHANGE_FNA_df.rename(columns={'Current Month': Current_Month1}, inplace=True)
FNAload_df = EXCHANGE_FNA_df[['Vehicle Line', Current_Month1]]

FNAload_df
Canadaload_df
DMload_df
FAPload_df
FSAload_df
FoEload_df
MEAload_df
Mexicoload_df
USAload_df


OP=py.load_workbook(output_file)

Sheet1 = OP.get_sheet_by_name('EXCHANGE US')
Sheet2 = OP.get_sheet_by_name('EXCHANGE FNA')
Sheet3 = OP.get_sheet_by_name('EXCHANGE Canada')
Sheet4 = OP.get_sheet_by_name('EXCHANGE DM')
Sheet5 = OP.get_sheet_by_name('EXCHANGE FAP')
Sheet6 = OP.get_sheet_by_name('EXCHANGE FSA')
Sheet7 = OP.get_sheet_by_name('EXCHANGE FoE')
Sheet8 = OP.get_sheet_by_name('EXCHANGE MEA')
Sheet9 = OP.get_sheet_by_name('EXCHANGE Mexico')


if Current_Month1 == "January":
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=3)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=3)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=3)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=3)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=3)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=3)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=3)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=3)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=3)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'February':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=5)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=5)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=5)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=5)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=5)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=5)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=5)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=5)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=5)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'March':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=7)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=7)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=7)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=7)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=7)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=7)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=7)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=7)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=7)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'April':
	    for i in range(9, 169):
            cell1 = Sheet1.cell(row=i, column=11)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=11)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=11)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=11)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=11)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=11)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=11)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=11)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=11)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'May':
	    for i in range(9, 169):
            cell1 = Sheet1.cell(row=i, column=13)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=13)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=13)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=13)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=13)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=13)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=13)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=13)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=13)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'June':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=15)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=15)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=15)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=15)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=15)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=15)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=15)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=15)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=15)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'July':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=19)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=19)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=19)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=19)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=19)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=19)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=19)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=19)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=19)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'August':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=21)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=21)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=21)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=21)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=21)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=21)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=21)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=21)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=21)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'September':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=23)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=23)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=23)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=23)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=23)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=23)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=23)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=23)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=23)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'October':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=27)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=27)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=27)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=27)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=27)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=27)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=27)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=27)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=27)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
elif Current_Month1 == 'November':
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=29)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=29)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=29)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=29)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=29)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=29)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=29)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=29)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=29)
            cell9.value = Mexicoload_df.iloc[i-7, 1]
else:
        for i in range(9, 170):
            cell1 = Sheet1.cell(row=i, column=31)
            cell1.value = USAload_df.iloc[i-7, 1]
            cell2 = Sheet2.cell(row=i, column=31)
            cell2.value = FNAload_df.iloc[i-7, 1]
            cell3 = Sheet3.cell(row=i, column=31)
            cell3.value = Canadaload_df.iloc[i-7, 1]
            cell4 = Sheet4.cell(row=i, column=31)
            cell4.value = DMload_df.iloc[i-7, 1]
            cell5 = Sheet5.cell(row=i, column=31)
            cell5.value = FAPload_df.iloc[i-7, 1]
            cell6 = Sheet6.cell(row=i, column=31)
            cell6.value = FSAload_df.iloc[i-7, 1]
            cell7 = Sheet7.cell(row=i, column=31)
            cell7.value = FoEload_df.iloc[i-7, 1]
            cell8 = Sheet8.cell(row=i, column=31)
            cell8.value = MEAload_df.iloc[i-7, 1]
            cell9 = Sheet9.cell(row=i, column=31)
            cell9.value = Mexicoload_df.iloc[i-7, 1]


OP.save(SD+"\Exchange.xlsx")

print("File is ready")